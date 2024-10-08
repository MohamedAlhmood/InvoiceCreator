import pandas as pd
import glob
from fpdf import FPDF
from pathlib import Path

from openpyxl.styles.builtins import total

filepaths = glob.glob('invoices/*.xlsx')

for filepath in filepaths:

    pdf = FPDF(orientation='P',unit='mm',format='A4')
    pdf.add_page()

    filename = Path(filepath).stem
    invoiceNr,date= filename.split('-')
    pdf.set_font(family="Times",size=16,style='B')
    pdf.cell(w=50,h=8,txt=f"Invoice #{invoiceNr}",ln=1)
    pdf.set_font(family="Times",size=16,style='B')
    pdf.cell(w=50,h=8,txt=f"Date: {date}",ln=1)
    pdf.set_font(family="Times",size=10,style='B')

    df = pd.read_excel(filepath,sheet_name='Sheet 1')
    columns=list(df.columns)
    columns1 = [item.replace('_'," ").title()for item in columns]
    pdf.cell(w=30, h=8, txt=columns1[0], border=1)
    pdf.cell(w=70, h=8, txt=columns1[1], border=1)
    pdf.cell(w=30, h=8, txt=columns1[2], border=1)
    pdf.cell(w=30, h=8, txt=columns1[3], border=1)
    pdf.cell(w=30, h=8, txt=columns1[4], border=1, ln=1)

    for index, row in df.iterrows():
        pdf.set_font("Times",size = 10)
        pdf.set_text_color(80,80,80)
        pdf.cell(w=30,h=8,txt=str(row["product_id"]),border=1)
        pdf.cell(w=70,h=8,txt=str(row['product_name']),border=1)
        pdf.cell(w=30,h=8,txt=str(row['amount_purchased']),border=1)
        pdf.cell(w=30,h=8,txt=str(row['price_per_unit']),border=1)
        pdf.cell(w=30,h=8,txt=str(row['total_price']),border=1,ln=1)
    totalSum=df['total_price'].sum()
    pdf.cell(w=30, h=8, txt='', border=1)
    pdf.cell(w=70, h=8, txt='', border=1)
    pdf.cell(w=30, h=8, txt='', border=1)
    pdf.cell(w=30, h=8, txt='', border=1)
    pdf.cell(w=30, h=8, txt=str(totalSum), border=1, ln=2)

    pdf.set_font("Times", size=10,style='B')
    pdf.cell(w=30, h=8, txt=f"The total price is {totalSum}", ln=1)

    pdf.set_font("Times", size=10)
    pdf.cell(w=30, h=8, txt="AlbanniProductions",ln=1)
    #pdf.image("murt.jpg",w=50)

    pdf.output(f'PDFS/{filename}.pdf')